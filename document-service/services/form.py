from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from copy import copy
import os
from schemas.employee import EmployeeVacationSendData


def fill_t7_form_with_openpyxl(employees: EmployeeVacationSendData, input_file=os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "document.xlsx"), output_file='отпуска_т7_filled.xlsx'):
    if employees is None:
        employees = []

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Файл шаблона {input_file} не найден!")

    # Загружаем книгу Excel
    wb = load_workbook(input_file)
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    start_row = 28
    col_offset = 2

    for i, employee in enumerate(employees.employees, start=start_row):
        for col in range(col_offset, col_offset + 10):
            if ws.cell(row=start_row - 1, column=col).has_style:
                ws.cell(row=i, column=col)._style = copy(ws.cell(row=start_row - 1, column=col)._style)

            ws.cell(row=i, column=col).border = thin_border

        ws.cell(row=i, column=col_offset, value=employee.department)  # C - Подразделение
        ws.cell(row=i, column=col_offset + 1, value=employee.position)  # D - Должность
        ws.cell(row=i, column=col_offset + 2, value=employee.name)  # E - ФИО
        ws.cell(row=i, column=col_offset + 4, value=employee.vacation_days)
        ws.cell(row=i, column=col_offset + 5, value=employee.advanced_days)
        ws.cell(row=i, column=col_offset + 6, value=int(employee.vacation_days) + int(employee.advanced_days))
        ws.cell(row=i, column=col_offset + 7, value=employee.planned_date.strftime('%d.%m.%Y'))

    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'

    wb.save(output_file)
    return output_file


# if __name__ == "__main__":
#     employees_data = [
#         {
#             'department': 'Отдел продаж',
#             'position': 'Менеджер по продажам',
#             'name': 'Иванов Иван Иванович',
#             'contract': 'КТ-001',
#             'vacation_days': 14,
#             "advanced_days": 2,
#             'planned_date': datetime.datetime.now()
#         },
#         {
#             'department': 'IT отдел',
#             'position': 'Старший разработчик',
#             'name': 'Петров Петр Петрович',
#             'contract': 'КТ-002',
#             'vacation_days': 21,
#             "advanced_days": 3,
#             'planned_date': datetime.datetime.now()
#         }
#     ]
#
#     try:
#         result_file = fill_t7_form_with_openpyxl(
#             input_file=os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "document.xlsx"),  # Ваш исходный файл
#             employees=employees_data,
#             output_file='отпуска_форма_т7_filled.xlsx'
#         )
#
#     except Exception as e:
#         print(f"Ошибка: {e}")
